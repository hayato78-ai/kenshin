#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
統合 Excel転記エンジン
労災二次検診 / 人間ドック 対応

exam_type に基づいて適切なトランスクライバーを選択
"""

import os
import sys
import json
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

try:
    import yaml
except ImportError:
    yaml = None

# ロギング設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class RosaiTranscriber:
    """労災二次検診 Excel転記エンジン（セル位置固定）"""

    # テンプレートパス
    TEMPLATE_PATH = "/Users/hytenhd_mac/Library/CloudStorage/GoogleDrive-buskenshin@cdmedical.jp/マイドライブ/40_労災二次検診/30_共通資料/81_結果入力/input/テンプレートファイル/kenshin_idheart.xlsx"
    SHEET_NAME = "template"

    # ============================================
    # セル位置定義（iD-Heart形式 固定）
    # ============================================
    CELLS = {
        # 基本情報
        'organization': 'B4',       # 事業所名
        'name': 'B5',               # 受診者名
        'gender': 'G5',             # 性別
        'birthdate': 'J5',          # 生年月日
        'age': 'O5',                # 年齢
        'visit_date': 'F18',        # 受診日

        # 問診
        'past_history': 'D14',      # 既往歴
        'subjective': 'D15',        # 自覚症状
        'objective': 'D16',         # 他覚症状

        # 超音波検査（所見=F列, 判定=H列）
        'cardiac_findings': 'F19',  # 心臓超音波 所見
        'cardiac_judgment': 'H19',  # 心臓超音波 判定
        'carotid_findings': 'F20',  # 頸動脈超音波 所見
        'carotid_judgment': 'H20',  # 頸動脈超音波 判定

        # 脂質（値=F列, H/L=D列, 判定=H列）
        'hdl_value': 'F21',
        'hdl_flag': 'D21',
        'hdl_judgment': 'H21',
        'ldl_value': 'F22',
        'ldl_flag': 'D22',
        'ldl_judgment': 'H22',
        'tg_value': 'F23',
        'tg_flag': 'D23',
        'tg_judgment': 'H23',

        # 糖代謝
        'fbs_value': 'F24',
        'fbs_flag': 'D24',
        'fbs_judgment': 'H24',
        'hba1c_value': 'F25',
        'hba1c_flag': 'D25',
        'hba1c_judgment': 'H25',

        # 腎機能
        'alb_cre_value': 'F26',     # アルブミン/クレアチニン
        'alb_cre_flag': 'D26',
        'alb_cre_judgment': 'H26',
        'alb_value': 'F27',         # アルブミン定量値
        'alb_flag': 'D27',
        'alb_judgment': 'H27',
        'cre_value': 'F28',         # クレアチニン(参考)
        'cre_flag': 'D28',
        'cre_judgment': 'H28',

        # 保健指導・総合所見（結合セル注意）
        'health_guidance': 'A31',   # 保健指導
        'general_findings': 'A34',  # 総合所見
    }

    def __init__(self):
        """初期化"""
        if not Path(self.TEMPLATE_PATH).exists():
            raise FileNotFoundError(f"テンプレートが見つかりません: {self.TEMPLATE_PATH}")
        logger.info("✅ RosaiTranscriber 初期化完了")

    def _safe_write(self, ws, cell: str, value: Any) -> bool:
        """セルに安全に書き込み（結合セル対応）"""
        if value is None or value == '':
            return False
        try:
            cell_obj = ws[cell]
            if isinstance(cell_obj, MergedCell):
                for merged_range in ws.merged_cells.ranges:
                    if cell_obj.coordinate in merged_range:
                        ws[merged_range.start_cell.coordinate] = value
                        return True
                return False
            else:
                ws[cell] = value
                return True
        except Exception as e:
            logger.warning(f"⚠️ セル書き込みエラー {cell}: {e}")
            return False

    def transcribe(self, data: Dict, output_path: str = None) -> Dict:
        """
        JSONデータをExcelに転記

        Args:
            data: GASからのリクエストJSON
            output_path: 出力パス（省略時は自動決定）

        Returns:
            {'success': bool, 'output_path': str, 'count': int}
        """
        try:
            request_id = data.get('request_id', 'UNKNOWN')
            logger.info(f"📝 転記開始: {request_id}")

            # テンプレート読み込み
            wb = load_workbook(self.TEMPLATE_PATH)
            ws = wb[self.SHEET_NAME]

            count = 0

            # ========== 基本情報 ==========
            case = data.get('case', {})
            patient = data.get('patient', {})

            count += self._safe_write(ws, self.CELLS['organization'], case.get('company_name'))
            count += self._safe_write(ws, self.CELLS['name'], patient.get('name'))
            count += self._safe_write(ws, self.CELLS['gender'], patient.get('gender'))
            # 日付は日本語形式に変換
            birthdate = self._format_date_jp(patient.get('birthdate'))
            count += self._safe_write(ws, self.CELLS['birthdate'], birthdate)
            count += self._safe_write(ws, self.CELLS['age'], patient.get('age'))
            # 受診日
            visit_date = self._format_date_jp(patient.get('visit_date') or case.get('exam_date'))
            if 'visit_date' in self.CELLS:
                count += self._safe_write(ws, self.CELLS['visit_date'], visit_date)

            # ========== 血液検査 ==========
            blood = data.get('blood_tests', {})
            # ネスト構造対応: lipid.hdl_c or hdl
            lipid = blood.get('lipid', {})
            glucose = blood.get('glucose', {})

            # HDL (row 21)
            hdl = blood.get('hdl') or blood.get('hdl_c') or lipid.get('hdl_c') or lipid.get('hdl') or {}
            if isinstance(hdl, dict):
                count += self._safe_write(ws, self.CELLS['hdl_value'], hdl.get('value'))
                count += self._safe_write(ws, self.CELLS['hdl_flag'], hdl.get('flag'))
                count += self._safe_write(ws, self.CELLS['hdl_judgment'], hdl.get('judgment'))

            # LDL (row 22)
            ldl = blood.get('ldl') or blood.get('ldl_c') or lipid.get('ldl_c') or lipid.get('ldl') or {}
            if isinstance(ldl, dict):
                count += self._safe_write(ws, self.CELLS['ldl_value'], ldl.get('value'))
                count += self._safe_write(ws, self.CELLS['ldl_flag'], ldl.get('flag'))
                count += self._safe_write(ws, self.CELLS['ldl_judgment'], ldl.get('judgment'))

            # TG (row 23)
            tg = blood.get('tg') or lipid.get('tg') or {}
            if isinstance(tg, dict):
                count += self._safe_write(ws, self.CELLS['tg_value'], tg.get('value'))
                count += self._safe_write(ws, self.CELLS['tg_flag'], tg.get('flag'))
                count += self._safe_write(ws, self.CELLS['tg_judgment'], tg.get('judgment'))

            # FBS (row 24)
            fbs = blood.get('fbs') or glucose.get('fbs') or {}
            if isinstance(fbs, dict):
                count += self._safe_write(ws, self.CELLS['fbs_value'], fbs.get('value'))
                count += self._safe_write(ws, self.CELLS['fbs_flag'], fbs.get('flag'))
                count += self._safe_write(ws, self.CELLS['fbs_judgment'], fbs.get('judgment'))

            # HbA1c (row 25)
            hba1c = blood.get('hba1c') or blood.get('hba1c_ngsp') or glucose.get('hba1c_ngsp') or glucose.get('hba1c') or {}
            if isinstance(hba1c, dict):
                count += self._safe_write(ws, self.CELLS['hba1c_value'], hba1c.get('value'))
                count += self._safe_write(ws, self.CELLS['hba1c_flag'], hba1c.get('flag'))
                count += self._safe_write(ws, self.CELLS['hba1c_judgment'], hba1c.get('judgment'))

            # ========== 超音波検査 ==========
            ultrasound = data.get('ultrasound', {})

            # 心臓超音波 (row 19)
            cardiac = ultrasound.get('cardiac', {})
            if cardiac:
                count += self._safe_write(ws, self.CELLS['cardiac_findings'], cardiac.get('findings'))
                count += self._safe_write(ws, self.CELLS['cardiac_judgment'], cardiac.get('judgment'))

            # 頸動脈超音波 (row 20)
            carotid = ultrasound.get('carotid', {})
            if carotid:
                count += self._safe_write(ws, self.CELLS['carotid_findings'], carotid.get('findings'))
                count += self._safe_write(ws, self.CELLS['carotid_judgment'], carotid.get('judgment'))

            # ========== 保健指導・総合所見 ==========
            guidance = data.get('guidance', {})

            # 保健指導 (A31)
            health_guidance = guidance.get('health_guidance') or guidance.get('text') or ''
            count += self._safe_write(ws, self.CELLS['health_guidance'], health_guidance)

            # 総合所見 (A34)
            doctor_findings = guidance.get('doctor_findings') or ''
            count += self._safe_write(ws, self.CELLS['general_findings'], doctor_findings)

            # ========== 出力パス決定 ==========
            if not output_path:
                output_config = data.get('output', {})
                folder_path = output_config.get('folder_path')

                if folder_path and Path(folder_path).exists():
                    output_dir = Path(folder_path)
                else:
                    # デフォルト: python/output
                    output_dir = Path(__file__).parent / 'output'

                output_dir.mkdir(parents=True, exist_ok=True)

                # ファイル名: 氏名_受診日.xlsx
                name = patient.get('name', 'unknown')
                visit_date = patient.get('visit_date') or datetime.now().strftime('%Y-%m-%d')
                filename = self._sanitize(f"{name}_{visit_date}.xlsx")
                output_path = output_dir / filename

            # 保存
            wb.save(output_path)
            logger.info(f"✅ 転記完了: {count}項目 → {output_path}")

            return {
                'success': True,
                'output_path': str(output_path),
                'count': count
            }

        except Exception as e:
            logger.error(f"❌ 転記エラー: {e}")
            return {'success': False, 'error': str(e)}

    def _sanitize(self, filename: str) -> str:
        """ファイル名をサニタイズ"""
        for char in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
            filename = filename.replace(char, '_')
        return filename

    def _format_date_jp(self, date_str: str) -> str:
        """日付を日本語形式に変換 (YYYY-MM-DD → YYYY年MM月DD日)"""
        if not date_str:
            return ''
        try:
            # YYYY-MM-DD形式をパース
            if '-' in str(date_str):
                parts = str(date_str).split('-')
                if len(parts) == 3:
                    year, month, day = parts
                    return f"{year}年{int(month)}月{int(day)}日"
            # 既に日本語形式ならそのまま
            if '年' in str(date_str):
                return str(date_str)
            return str(date_str)
        except:
            return str(date_str)


# ============================================================
# 人間ドック用トランスクライバー
# ============================================================
class HumanDockTranscriber:
    """人間ドック Excel転記エンジン（BML CSV → template.xlsm）"""

    def __init__(self, settings_path: Path = None):
        """初期化"""
        self.settings = self._load_settings(settings_path)
        exam_config = self.settings.get('exam_types', {}).get('HUMAN_DOCK', {})

        self.template_path = Path(exam_config.get('template_path', ''))
        self.mapping_path = Path(exam_config.get('mapping_path', ''))
        self.output_dir = Path(self.settings.get('output_dir', './output'))
        self.sheet_cache = {}  # シート名 → wsオブジェクトのキャッシュ

        if not self.template_path.exists():
            raise FileNotFoundError(f"テンプレートが見つかりません: {self.template_path}")
        if not self.mapping_path.exists():
            raise FileNotFoundError(f"マッピングファイルが見つかりません: {self.mapping_path}")

        # マッピング読み込み
        with open(self.mapping_path, 'r', encoding='utf-8') as f:
            self.mapping = json.load(f)

        # マッピングをフラット化（BMLコード → セル情報）
        self.flat_mapping = self._flatten_mapping()

        # 判定エンジン初期化
        sys.path.insert(0, str(Path(__file__).parent))
        try:
            from common import JudgmentEngine, GENDER_CODE_TO_INTERNAL
            self.judgment_engine = JudgmentEngine(
                self.mapping.get('judgment_criteria', {}).get('items', {})
            )
            self.GENDER_CODE_TO_INTERNAL = GENDER_CODE_TO_INTERNAL
        except ImportError:
            logger.warning("common モジュールをインポートできません。判定なしで実行")
            self.judgment_engine = None
            self.GENDER_CODE_TO_INTERNAL = {'1': 'M', '2': 'F'}

        logger.info("✅ HumanDockTranscriber 初期化完了")

    def _load_settings(self, settings_path: Path = None) -> dict:
        """settings.yaml を読み込む"""
        if settings_path is None:
            settings_path = Path(__file__).parent / 'settings.yaml'

        if settings_path.exists() and yaml:
            with open(settings_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f)
        return {}

    def _flatten_mapping(self) -> Dict[str, Dict]:
        """ネストしたマッピングをBMLコード→セル情報のフラットマップに変換"""
        flat = {}
        items = self.mapping.get('test_items', {}).get('items', {})
        for category, category_items in items.items():
            # コメントカテゴリはスキップ
            if category.startswith('_comment'):
                continue
            if isinstance(category_items, dict):
                for code, spec in category_items.items():
                    # コメントキーはスキップ
                    if code.startswith('_comment'):
                        continue
                    if isinstance(spec, dict) and 'value_cell' in spec:
                        flat[code] = spec
        logger.info(f"  マッピング読込: {len(flat)}項目")
        return flat

    def _get_sheet(self, wb, sheet_name: str):
        """シートを取得（キャッシュ使用）"""
        if sheet_name not in self.sheet_cache:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"シート '{sheet_name}' が見つかりません。利用可能: {wb.sheetnames}")
            self.sheet_cache[sheet_name] = wb[sheet_name]
        return self.sheet_cache[sheet_name]

    def transcribe_from_csv(self, csv_path: str, output_path: str = None) -> Dict:
        """
        BML CSVファイルからExcelに転記

        Args:
            csv_path: 入力CSVパス
            output_path: 出力パス（省略時は自動決定）

        Returns:
            {'success': bool, 'output_path': str, 'count': int}
        """
        try:
            # CSVパース
            sys.path.insert(0, str(Path(__file__).parent))
            from common import BMLResultParser
            parser = BMLResultParser()
            results = parser.parse(Path(csv_path))

            if not results:
                return {'success': False, 'error': 'CSVに患者データが見つかりません'}

            # 最初の患者を処理（複数患者は将来対応）
            patient_data = results[0]
            patient_info = patient_data['patient_info']
            test_results = patient_data['test_results']

            # テンプレート読み込み（VBA保持）
            wb = load_workbook(self.template_path, keep_vba=True)
            self.sheet_cache = {}  # キャッシュクリア

            count = 0

            # 性別判定
            gender_code = patient_info.get('gender', '')
            gender = self.GENDER_CODE_TO_INTERNAL.get(gender_code, 'M')

            # 患者情報転記
            count += self._transfer_patient_info(wb, patient_info, gender)

            # 検査結果転記
            count += self._transfer_test_results(wb, test_results, gender)

            # 出力パス決定
            if not output_path:
                self.output_dir.mkdir(parents=True, exist_ok=True)
                request_id = patient_info.get('request_id', 'unknown')
                exam_date = patient_info.get('exam_date', datetime.now().strftime('%Y%m%d'))
                output_path = self.output_dir / f"result_{exam_date}_{request_id}.xlsm"

            # 保存
            wb.save(output_path)
            logger.info(f"✅ 転記完了: {count}項目 → {output_path}")

            return {
                'success': True,
                'output_path': str(output_path),
                'count': count
            }

        except Exception as e:
            logger.error(f"❌ 転記エラー: {e}")
            import traceback
            traceback.print_exc()
            return {'success': False, 'error': str(e)}

    def transcribe(self, data: Dict, output_path: str = None) -> Dict:
        """
        JSONリクエストからExcelに転記（GAS連携用）

        Args:
            data: GASからのリクエストJSON
            output_path: 出力パス

        Returns:
            {'success': bool, 'output_path': str, 'count': int}
        """
        # CSV パス指定の場合
        csv_path = data.get('csv_path')
        if csv_path:
            return self.transcribe_from_csv(csv_path, output_path)

        # JSON直接データの場合（GAS方式）
        patient_info = data.get('patient_info')
        test_results_dict = data.get('test_results')

        if patient_info and isinstance(patient_info, dict) and len(patient_info) > 0:
            try:
                logger.info(f"📋 JSON直接処理モード: {patient_info.get('name', 'UNKNOWN')}")

                # test_results を辞書形式からリスト形式に変換
                # {"0000301": {"value": 4950, ...}} → [{"code": "0000301", "value": 4950, ...}]
                test_results_list = []
                if test_results_dict and isinstance(test_results_dict, dict):
                    for code, values in test_results_dict.items():
                        if isinstance(values, dict):
                            item = {'code': code, **values}
                        else:
                            item = {'code': code, 'value': values}
                        test_results_list.append(item)
                    logger.info(f"  検査結果: {len(test_results_list)}項目")

                # テンプレート読み込み（VBA保持）
                wb = load_workbook(self.template_path, keep_vba=True)
                self.sheet_cache = {}  # キャッシュクリア

                count = 0

                # 性別判定（GASからの形式に対応）
                gender_raw = patient_info.get('gender', '')
                if gender_raw in ['男', 'M', '1']:
                    gender = 'M'
                elif gender_raw in ['女', 'F', '2']:
                    gender = 'F'
                else:
                    gender = self.GENDER_CODE_TO_INTERNAL.get(gender_raw, 'M')

                # 患者情報転記
                count += self._transfer_patient_info(wb, patient_info, gender)

                # 検査結果転記
                count += self._transfer_test_results(wb, test_results_list, gender)

                # 出力パス決定
                if not output_path:
                    self.output_dir.mkdir(parents=True, exist_ok=True)
                    request_id = data.get('request_id', 'unknown')
                    exam_date = patient_info.get('examDate', datetime.now().strftime('%Y%m%d'))
                    # 日付フォーマットを調整（2025/12/24 → 20251224）
                    if '/' in str(exam_date):
                        exam_date = exam_date.replace('/', '')
                    output_path = self.output_dir / f"result_{exam_date}_{request_id}.xlsm"

                # 保存
                wb.save(output_path)
                logger.info(f"✅ 転記完了: {count}項目 → {output_path}")

                return {
                    'success': True,
                    'output_path': str(output_path),
                    'count': count
                }

            except Exception as e:
                logger.error(f"❌ JSON直接処理エラー: {e}")
                import traceback
                traceback.print_exc()
                return {'success': False, 'error': str(e)}

        # データがない場合はエラー
        return {'success': False, 'error': 'csv_path または patient_info が指定されていません'}

    def _transfer_patient_info(self, wb, patient_info: Dict, gender: str) -> int:
        """患者基本情報を転記（複数シート対応）"""
        count = 0
        patient_map = self.mapping.get('patient_info', {})

        # マッピングに患者情報定義がある場合
        for field, spec in patient_map.items():
            if isinstance(spec, dict) and 'cell' in spec:
                sheet_name = spec.get('sheet', '１ページ')
                try:
                    ws = self._get_sheet(wb, sheet_name)
                    value = patient_info.get(field)
                    if value:
                        ws[spec['cell']] = value
                        count += 1
                except Exception as e:
                    logger.warning(f"⚠️ 患者情報転記エラー {field}: {e}")

        logger.info(f"  患者情報: {count}項目転記 (request_id={patient_info.get('request_id')})")
        return count

    def _transfer_test_results(self, wb, test_results: list, gender: str) -> int:
        """検査結果を複数シートに転記"""
        count = 0

        for result in test_results:
            code = result.get('code') or result.get('item_code')
            if not code:
                continue

            # フラット化されたマッピングから取得
            if code not in self.flat_mapping:
                continue

            spec = self.flat_mapping[code]
            sheet_name = spec.get('sheet', '４ページ')

            try:
                ws = self._get_sheet(wb, sheet_name)

                # 値を転記（M列）
                value_cell = spec.get('value_cell')
                if value_cell:
                    raw_value = result.get('value')
                    if raw_value is not None and raw_value != '':
                        try:
                            numeric_value = float(raw_value)
                            ws[value_cell] = numeric_value
                        except (ValueError, TypeError):
                            ws[value_cell] = raw_value
                        count += 1
                        logger.debug(f"  {code} → {sheet_name}!{value_cell}: {raw_value}")

                # 判定を転記（K列）
                judgment_cell = spec.get('judgment_cell')
                if judgment_cell:
                    judgment = result.get('judgment')
                    if judgment:
                        ws[judgment_cell] = judgment
                        count += 1
                    elif self.judgment_engine:
                        # 判定エンジンで自動判定
                        auto_judgment = self.judgment_engine.judge_by_code(
                            code, result.get('value'), result.get('flag', ''), gender
                        )
                        if auto_judgment:
                            ws[judgment_cell] = auto_judgment
                            count += 1

                # H/Lフラグ（flag_cellがある場合）
                flag_cell = spec.get('flag_cell')
                if flag_cell and result.get('flag'):
                    ws[flag_cell] = result['flag']
                    count += 1

            except Exception as e:
                logger.warning(f"⚠️ 転記エラー {code}: {e}")

        logger.info(f"  検査結果: {count}項目転記")
        return count


# ============================================================
# DriveWatcher用エントリーポイント
# ============================================================
def process_export_request(request_data: Dict) -> Dict:
    """
    DriveWatcherから呼び出されるエントリーポイント
    exam_type に基づいて適切なトランスクライバーを選択
    """
    exam_type = request_data.get('exam_type', 'ROSAI_SECONDARY')

    if exam_type == 'HUMAN_DOCK':
        logger.info("📋 人間ドック モードで処理")
        transcriber = HumanDockTranscriber()
        return transcriber.transcribe(request_data)
    else:
        logger.info("📋 労災二次検診 モードで処理")
        transcriber = RosaiTranscriber()
        return transcriber.transcribe(request_data)


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='統合 Excel転記エンジン')
    parser.add_argument('json_file', nargs='?', help='入力JSONファイル')
    parser.add_argument('--csv', help='BML CSVファイル（人間ドック直接実行用）')
    parser.add_argument('--output', help='出力パス')
    parser.add_argument('--type', choices=['ROSAI_SECONDARY', 'HUMAN_DOCK'],
                        default='ROSAI_SECONDARY', help='検査種別')
    parser.add_argument('--watch', action='store_true',
                        help='監視モード: pendingフォルダを監視して自動処理')
    parser.add_argument('--settings', default='settings.yaml',
                        help='設定ファイルパス（監視モード用）')

    args = parser.parse_args()

    # 監視モード
    if args.watch:
        from drive_watcher import DriveWatcher
        print("🚀 監視モード起動")
        watcher = DriveWatcher.from_settings(args.settings, process_export_request)
        watcher.start()
        sys.exit(0)  # 監視終了後は正常終了
    # CSVモード（人間ドック直接実行）
    elif args.csv:
        transcriber = HumanDockTranscriber()
        result = transcriber.transcribe_from_csv(args.csv, args.output)
    # JSONモード
    elif args.json_file:
        with open(args.json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        data['exam_type'] = args.type
        result = process_export_request(data)
    else:
        parser.print_help()
        sys.exit(1)

    if result['success']:
        print(f"✅ 成功: {result['output_path']}")
        print(f"   転記項目数: {result['count']}")
    else:
        print(f"❌ 失敗: {result['error']}")
        sys.exit(1)
