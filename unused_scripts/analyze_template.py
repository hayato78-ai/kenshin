#!/usr/bin/env python3
"""
Excelテンプレートの構造を解析するスクリプト
"""
import json
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxlをインストールしてください: pip install openpyxl")
    exit(1)

# パス設定
BASE_DIR = Path(__file__).parent.parent
TEMPLATE_PATH = BASE_DIR / "結果入力テンプレ" / "20250908_フォーマット健康診断結果表(サンプル値あり).xlsm"

def analyze_template():
    """テンプレートの構造を解析"""
    wb = load_workbook(TEMPLATE_PATH, read_only=True, data_only=True)

    print("=" * 60)
    print("シート一覧:")
    print("=" * 60)
    for i, sheet_name in enumerate(wb.sheetnames):
        print(f"  {i}: {sheet_name}")

    # 項目評価シートを解析
    print("\n" + "=" * 60)
    print("項目評価シートの内容:")
    print("=" * 60)

    ws = wb["項目評価"]

    # 値のあるセルを抽出
    data_cells = []
    for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=40):
        for cell in row:
            if cell.value is not None:
                data_cells.append({
                    "cell": cell.coordinate,
                    "row": cell.row,
                    "col": cell.column,
                    "value": str(cell.value)[:50]  # 長い値は切り詰め
                })

    # セル位置と値を表示
    for item in data_cells[:100]:  # 最初の100セル
        print(f"  {item['cell']:6} | {item['value']}")

    # 判定マスタシートを解析
    print("\n" + "=" * 60)
    print("判定マスタシートの内容:")
    print("=" * 60)

    ws_master = wb["判定マスタ"]

    # ヘッダー行を取得
    headers = []
    for cell in ws_master[3]:
        if cell.value:
            headers.append((cell.column, cell.value))

    print("ヘッダー:")
    for col, val in headers:
        print(f"  列{col}: {val}")

    # 判定データを取得
    print("\n判定データ:")
    for row in ws_master.iter_rows(min_row=4, max_row=35, min_col=1, max_col=20):
        values = [str(cell.value) if cell.value else "" for cell in row]
        if values[0]:  # 項目名がある行のみ
            print(f"  {values[0][:20]:20} | A:{values[8]}-{values[10]} | B:{values[11]}-{values[13]} | C:{values[14]}-{values[16]} | D:{values[17]}-{values[19]}")

    wb.close()

if __name__ == "__main__":
    analyze_template()
