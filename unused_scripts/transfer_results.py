#!/usr/bin/env python3
"""
健診結果転記スクリプト
検査結果CSV → Excelテンプレート への自動転記

対応CSVフォーマット:
  - BML結果CSV (09XXXXXX_O6319101.csv形式)

使用方法:
    python transfer_results.py <入力CSV> [出力Excel]
"""

import sys
import json
from pathlib import Path
from typing import Dict, Any, List
import shutil

try:
    import yaml
except ImportError:
    yaml = None

try:
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook
except ImportError:
    print("ERROR: openpyxlをインストールしてください")
    print("  pip install openpyxl")
    sys.exit(1)

# 共通モジュールのインポート
from common import BMLResultParser, JudgmentEngine, GENDER_CODE_TO_INTERNAL


class ResultTransfer:
    """検査結果転記クラス"""

    def __init__(self, mapping_path: Path):
        """
        Args:
            mapping_path: マッピング設定JSONファイルパス
        """
        with open(mapping_path, "r", encoding="utf-8") as f:
            self.mapping = json.load(f)

        self.judgment_engine = JudgmentEngine(
            self.mapping.get("judgment_criteria", {}).get("items", {})
        )
        self.parser = BMLResultParser()

    def process_csv(self, csv_path: Path) -> List[Dict]:
        """CSVを解析して全患者の結果を取得"""
        return self.parser.parse(csv_path)

    def transfer_to_excel(
        self,
        template_path: Path,
        output_path: Path,
        patient_data: Dict
    ) -> Dict[str, Any]:
        """
        1患者分のデータをExcelに転記

        Args:
            template_path: テンプレートExcelファイルパス
            output_path: 出力Excelファイルパス
            patient_data: 患者データ辞書

        Returns:
            転記統計情報
        """
        wb = self._load_workbook(template_path, output_path)
        stats = {"transferred": 0, "skipped": 0, "errors": []}

        patient_info = patient_data['patient_info']
        test_results = patient_data['test_results']

        # 性別を判定
        gender_code = patient_info.get('gender', '')
        gender = GENDER_CODE_TO_INTERNAL.get(gender_code, "M")

        # シート取得
        sheet_name = self.mapping.get("test_items", {}).get("sheet", "項目評価")
        try:
            ws = wb[sheet_name]
        except KeyError:
            stats["errors"].append(f"シート '{sheet_name}' が見つかりません")
            wb.close()
            return stats

        # 患者基本情報の転記
        self._transfer_patient_info(ws, patient_info, gender, stats)

        # 検査結果の転記
        self._transfer_test_results(ws, test_results, gender, stats)

        # 保存
        wb.save(output_path)
        wb.close()

        return stats

    def _load_workbook(self, template_path: Path, output_path: Path) -> Workbook:
        """Excelワークブックを読み込み"""
        output_ext = output_path.suffix.lower()
        template_ext = template_path.suffix.lower()

        if template_ext == ".xlsm" and output_ext == ".xlsx":
            return load_workbook(template_path, keep_vba=False)
        elif template_ext == ".xlsm" and output_ext == ".xlsm":
            shutil.copy(template_path, output_path)
            return load_workbook(output_path, keep_vba=True)
        else:
            return load_workbook(template_path, keep_vba=False)

    def _transfer_patient_info(self, ws, patient_info: Dict, gender: str, stats: Dict):
        """患者基本情報を転記"""
        try:
            # 受診日
            exam_date = patient_info.get('exam_date', '')
            if exam_date and len(exam_date) == 8:
                formatted_date = f"{exam_date[:4]}/{exam_date[4:6]}/{exam_date[6:]}"
                ws['AG5'] = formatted_date

            # 依頼ID
            ws['G8'] = patient_info.get('request_id', '')

            # 性別
            ws['S9'] = "男" if gender == "M" else "女"

            stats["transferred"] += 3
        except Exception as e:
            stats["errors"].append(f"患者情報転記エラー: {e}")

    def _transfer_test_results(self, ws, test_results: List[Dict], gender: str, stats: Dict):
        """検査結果を転記"""
        test_item_map = self.mapping.get("test_items", {}).get("items", {})

        for result in test_results:
            code = result['code']

            if code not in test_item_map:
                stats["skipped"] += 1
                continue

            item_spec = test_item_map[code]

            try:
                value_cell = item_spec.get("value_cell")
                judgment_cell = item_spec.get("judgment_cell")

                # 結果値を転記
                if value_cell:
                    raw_value = result['value']
                    try:
                        numeric_value = float(raw_value)
                        ws[value_cell] = numeric_value
                    except (ValueError, TypeError):
                        ws[value_cell] = raw_value

                # 判定を転記
                if judgment_cell:
                    judgment = self.judgment_engine.judge_by_code(
                        code, result['value'], result['flag'], gender
                    )
                    ws[judgment_cell] = judgment

                stats["transferred"] += 1

            except Exception as e:
                stats["errors"].append(f"検査項目 {code}: {e}")
                stats["skipped"] += 1


def load_settings(base_dir: Path) -> dict:
    """settings.yamlから設定を読み込む"""
    settings_path = base_dir / "python" / "settings.yaml"
    if settings_path.exists() and yaml:
        with open(settings_path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f)
    return {}


def main():
    if len(sys.argv) < 2:
        print("使用方法: python transfer_results.py <入力CSV> [出力Excel]")
        print("")
        print("例:")
        print("  python transfer_results.py 09090947_O6319101.csv")
        print("  python transfer_results.py result.csv output.xlsx")
        sys.exit(1)

    script_dir = Path(__file__).parent
    base_dir = script_dir.parent

    input_csv = Path(sys.argv[1])
    if not input_csv.is_absolute():
        if not input_csv.exists():
            input_csv = base_dir / "01_csv" / sys.argv[1]

    if not input_csv.exists():
        print(f"ERROR: 入力ファイルが見つかりません: {input_csv}")
        sys.exit(1)

    # settings.yamlから設定を読み込み
    settings = load_settings(base_dir)
    exam_config = settings.get("exam_types", {}).get("HUMAN_DOCK", {})

    # マッピングファイルパス（settings優先）
    if exam_config.get("mapping_path"):
        mapping_path = Path(exam_config["mapping_path"])
    else:
        mapping_path = base_dir / "設計書_設定ファイル" / "mapping.json"

    # テンプレートパス（settings優先）
    if exam_config.get("template_path"):
        template_path = Path(exam_config["template_path"])
    else:
        template_path = base_dir / "結果入力テンプレ" / "テンプレフォルダ" / "template.xlsm"

    if not mapping_path.exists():
        print(f"ERROR: マッピングファイルが見つかりません: {mapping_path}")
        sys.exit(1)

    if not template_path.exists():
        print(f"ERROR: テンプレートファイルが見つかりません: {template_path}")
        sys.exit(1)

    print("=" * 60)
    print("健診結果転記システム (BML形式対応)")
    print("=" * 60)
    print(f"入力CSV: {input_csv.name}")
    print("")

    transfer = ResultTransfer(mapping_path)

    print("CSVを解析中...")
    all_results = transfer.process_csv(input_csv)
    print(f"  検出患者数: {len(all_results)} 名")
    print("")

    output_dir = base_dir / "02_出力フォルダ"
    output_dir.mkdir(parents=True, exist_ok=True)

    total_stats = {"transferred": 0, "skipped": 0, "files": 0}

    for i, patient_data in enumerate(all_results):
        patient_info = patient_data['patient_info']
        request_id = patient_info.get('request_id', f'unknown_{i}')
        exam_date = patient_info.get('exam_date', 'nodate')

        # 出力ファイル名を生成
        if len(sys.argv) >= 3 and len(all_results) == 1:
            output_path = Path(sys.argv[2])
            if not output_path.is_absolute():
                output_path = output_dir / sys.argv[2]
        else:
            output_path = output_dir / f"result_{exam_date}_{request_id}.xlsx"

        print(f"[{i+1}/{len(all_results)}] 依頼ID: {request_id}")
        print(f"  検査項目数: {len(patient_data['test_results'])}")

        stats = transfer.transfer_to_excel(template_path, output_path, patient_data)

        print(f"  転記成功: {stats['transferred']} 項目")
        print(f"  スキップ: {stats['skipped']} 項目")
        print(f"  出力: {output_path.name}")

        if stats.get("errors"):
            for err in stats["errors"][:3]:
                print(f"  エラー: {err}")

        print("")

        total_stats["transferred"] += stats["transferred"]
        total_stats["skipped"] += stats["skipped"]
        total_stats["files"] += 1

    print("=" * 60)
    print("処理完了")
    print("=" * 60)
    print(f"  出力ファイル数: {total_stats['files']}")
    print(f"  総転記項目数: {total_stats['transferred']}")
    print(f"  総スキップ数: {total_stats['skipped']}")
    print(f"  出力先: {output_dir}")


if __name__ == "__main__":
    main()
