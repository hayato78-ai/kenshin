#!/usr/bin/env python3
"""
健診結果転記スクリプト (template.xlsm用)
検査結果CSV → template.xlsm への自動転記

対応CSVフォーマット:
  - BML結果CSV (09XXXXXX_O6319101.csv形式)

使用方法:
    python transfer_template.py <入力CSV> [出力Excel]
"""

import sys
import json
from pathlib import Path
from typing import Dict, Any, List
import shutil

try:
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook
except ImportError:
    print("ERROR: openpyxlをインストールしてください")
    print("  pip install openpyxl")
    sys.exit(1)

# 共通モジュールのインポート
from common import BMLResultParser, JudgmentEngine, GENDER_CODE_TO_INTERNAL


class TemplateTransfer:
    """template.xlsm用の転記クラス"""

    def __init__(self, mapping_path: Path):
        """
        Args:
            mapping_path: マッピング設定JSONファイルパス
        """
        with open(mapping_path, "r", encoding="utf-8") as f:
            self.mapping = json.load(f)

        self.judgment_engine = JudgmentEngine(
            self.mapping.get("judgment_criteria", {}).get("items", {})
        )
        self.parser = BMLResultParser()

    def process_csv(self, csv_path: Path) -> List[Dict]:
        """CSVを解析して全患者の結果を取得"""
        return self.parser.parse(csv_path)

    def transfer_to_excel(
        self,
        template_path: Path,
        output_path: Path,
        patient_data: Dict
    ) -> Dict[str, Any]:
        """
        1患者分のデータをExcelに転記（複数シート対応）

        Args:
            template_path: テンプレートExcelファイルパス
            output_path: 出力Excelファイルパス
            patient_data: 患者データ辞書

        Returns:
            転記統計情報
        """
        wb = self._load_workbook(template_path, output_path)
        stats = {"transferred": 0, "skipped": 0, "errors": [], "details": []}

        patient_info = patient_data['patient_info']
        test_results = patient_data['test_results']

        # 性別を内部表現に変換
        gender_code = patient_info.get('gender', '')
        gender = GENDER_CODE_TO_INTERNAL.get(gender_code, "M")

        # 基本情報の転記
        self._transfer_patient_info(wb, patient_info, gender, stats)

        # 検査結果の転記（値と判定）
        self._transfer_test_results(wb, test_results, gender, stats)

        # 保存
        wb.save(output_path)
        wb.close()

        return stats

    def _load_workbook(self, template_path: Path, output_path: Path) -> Workbook:
        """Excelワークブックを読み込み"""
        output_ext = output_path.suffix.lower()
        template_ext = template_path.suffix.lower()

        if template_ext == ".xlsm" and output_ext == ".xlsx":
            return load_workbook(template_path, keep_vba=False)
        elif template_ext == ".xlsm" and output_ext == ".xlsm":
            shutil.copy(template_path, output_path)
            return load_workbook(output_path, keep_vba=True)
        else:
            return load_workbook(template_path, keep_vba=False)

    def _transfer_patient_info(
        self,
        wb: Workbook,
        patient_info: Dict,
        gender: str,
        stats: Dict
    ):
        """基本情報を1ページに転記"""
        patient_mapping = self.mapping.get("patient_info", {})
        sheet_name = patient_mapping.get("sheet", "1ページ")

        try:
            ws = wb[sheet_name]
        except KeyError:
            stats["errors"].append(f"シート '{sheet_name}' が見つかりません")
            return

        fields = patient_mapping.get("fields", {})

        # 受診No
        self._transfer_field(ws, fields, "request_id", patient_info.get('request_id', ''), stats)

        # 受診日（フォーマット変換）
        exam_date = patient_info.get('exam_date', '')
        if exam_date and len(exam_date) == 8:
            formatted = f"{exam_date[:4]}/{exam_date[4:6]}/{exam_date[6:]}"
            self._transfer_field(ws, fields, "exam_date", formatted, stats)

        # 性別（変換適用）
        if "gender" in fields:
            cell = fields["gender"].get("cell")
            if cell:
                transforms = self.mapping.get("transforms", {}).get("gender", {})
                gender_val = patient_info.get('gender', '')
                ws[cell] = transforms.get(gender_val, gender_val)
                stats["transferred"] += 1
                stats["details"].append(f"性別 → {cell}")

    def _transfer_field(
        self,
        ws,
        fields: Dict,
        field_name: str,
        value: Any,
        stats: Dict
    ):
        """単一フィールドを転記"""
        if field_name not in fields:
            return

        cell = fields[field_name].get("cell")
        if cell and value:
            ws[cell] = value
            stats["transferred"] += 1
            stats["details"].append(f"{field_name} → {cell}")

    def _transfer_test_results(
        self,
        wb: Workbook,
        test_results: List[Dict],
        gender: str,
        stats: Dict
    ):
        """検査結果を各シートに転記（値と判定）"""
        sheets_config = self.mapping.get("test_items", {}).get("sheets", {})

        # 全シートの全項目マッピングを統合
        all_items = {}
        for sheet_name, sheet_config in sheets_config.items():
            items = sheet_config.get("items", {})
            for code, spec in items.items():
                all_items[code] = {**spec, "sheet": sheet_name}

        for result in test_results:
            code = result['code']

            if code not in all_items:
                stats["skipped"] += 1
                continue

            item_spec = all_items[code]
            sheet_name = item_spec["sheet"]

            try:
                ws = wb[sheet_name]
            except KeyError:
                stats["errors"].append(f"シート '{sheet_name}' が見つかりません")
                stats["skipped"] += 1
                continue

            self._transfer_single_result(ws, result, item_spec, gender, stats, sheet_name)

    def _transfer_single_result(
        self,
        ws,
        result: Dict,
        item_spec: Dict,
        gender: str,
        stats: Dict,
        sheet_name: str
    ):
        """単一の検査結果を転記"""
        code = result['code']
        raw_value = result['value']
        flag = result['flag']

        try:
            # 結果値を転記
            value_cell = item_spec.get("value_cell")
            if value_cell:
                try:
                    numeric_value = float(raw_value)
                    ws[value_cell] = numeric_value
                except (ValueError, TypeError):
                    ws[value_cell] = raw_value

                stats["transferred"] += 1
                stats["details"].append(
                    f"{item_spec.get('name', code)} ({raw_value}) → {sheet_name}!{value_cell}"
                )

            # 判定を転記
            judgment_cell = item_spec.get("judgment_cell")
            if judgment_cell:
                judgment = self.judgment_engine.judge_by_code(code, raw_value, flag, gender)
                if judgment:
                    ws[judgment_cell] = judgment
                    stats["details"].append(
                        f"{item_spec.get('name', code)} 判定({judgment}) → {sheet_name}!{judgment_cell}"
                    )

        except Exception as e:
            stats["errors"].append(f"検査項目 {code}: {e}")
            stats["skipped"] += 1


def main():
    if len(sys.argv) < 2:
        print("使用方法: python transfer_template.py <入力CSV> [出力Excel]")
        print("")
        print("例:")
        print("  python transfer_template.py 09090947_O6319101.csv")
        print("  python transfer_template.py result.csv output.xlsx")
        sys.exit(1)

    script_dir = Path(__file__).parent
    base_dir = script_dir.parent

    input_csv = Path(sys.argv[1])
    if not input_csv.is_absolute():
        if not input_csv.exists():
            input_csv = base_dir / "01_csv" / sys.argv[1]

    if not input_csv.exists():
        print(f"ERROR: 入力ファイルが見つかりません: {input_csv}")
        sys.exit(1)

    mapping_path = base_dir / "設計書_設定ファイル" / "mapping_template.json"
    template_path = base_dir / "結果入力テンプレ" / "template.xlsm"

    if not mapping_path.exists():
        print(f"ERROR: マッピングファイルが見つかりません: {mapping_path}")
        sys.exit(1)

    if not template_path.exists():
        print(f"ERROR: テンプレートファイルが見つかりません: {template_path}")
        sys.exit(1)

    print("=" * 60)
    print("健診結果転記システム (template.xlsm用)")
    print("=" * 60)
    print(f"入力CSV: {input_csv.name}")
    print(f"テンプレート: {template_path.name}")
    print("")

    transfer = TemplateTransfer(mapping_path)

    print("CSVを解析中...")
    all_results = transfer.process_csv(input_csv)
    print(f"  検出患者数: {len(all_results)} 名")
    print("")

    output_dir = base_dir / "02_出力フォルダ"
    output_dir.mkdir(parents=True, exist_ok=True)

    total_stats = {"transferred": 0, "skipped": 0, "files": 0}

    for i, patient_data in enumerate(all_results):
        patient_info = patient_data['patient_info']
        request_id = patient_info.get('request_id', f'unknown_{i}')
        exam_date = patient_info.get('exam_date', 'nodate')

        if len(sys.argv) >= 3 and len(all_results) == 1:
            output_path = Path(sys.argv[2])
            if not output_path.is_absolute():
                output_path = output_dir / sys.argv[2]
        else:
            output_path = output_dir / f"result_{exam_date}_{request_id}.xlsx"

        print(f"[{i+1}/{len(all_results)}] 依頼ID: {request_id}")
        print(f"  検査項目数: {len(patient_data['test_results'])}")

        stats = transfer.transfer_to_excel(template_path, output_path, patient_data)

        print(f"  転記成功: {stats['transferred']} 項目")
        print(f"  スキップ: {stats['skipped']} 項目")
        print(f"  出力: {output_path.name}")

        # 転記詳細を表示
        if stats.get("details"):
            print("  転記内容:")
            for detail in stats["details"][:10]:
                print(f"    - {detail}")
            if len(stats["details"]) > 10:
                print(f"    ... 他 {len(stats['details']) - 10} 項目")

        if stats.get("errors"):
            for err in stats["errors"][:3]:
                print(f"  エラー: {err}")

        print("")

        total_stats["transferred"] += stats["transferred"]
        total_stats["skipped"] += stats["skipped"]
        total_stats["files"] += 1

    print("=" * 60)
    print("処理完了")
    print("=" * 60)
    print(f"  出力ファイル数: {total_stats['files']}")
    print(f"  総転記項目数: {total_stats['transferred']}")
    print(f"  総スキップ数: {total_stats['skipped']}")
    print(f"  出力先: {output_dir}")


if __name__ == "__main__":
    main()
